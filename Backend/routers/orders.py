# Backend/routers/orders.py
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Request
from sqlalchemy.orm import Session, joinedload  # 🌟 FIX: Ditambahkan joinedload agar tidak NameError
from pydantic import BaseModel
from typing import Optional
import json

import openpyxl
import models
import schemas
from dependencies import get_db, get_settings, get_current_user, require_role
from services import order_import_service
from services.order_service import (
    OrderService,
    OrderServiceError,
    OrderNotFoundError,
    OrderValidationError,
)
from services.order_fsm import validate_status_transition  
from main import limiter  

import pandas as pd
import logging

logger = logging.getLogger(__name__)

# =======================================================
# ROUTER CONFIG
# =======================================================
router = APIRouter(prefix="/api", tags=["Orders & Delivery"])

class TimeUpdateRequest(BaseModel):
    jam_maksimal: str  

class CoordinateUpdateRequest(BaseModel):
    latitude: float
    longitude: float
    kode_customer: str
    nama_customer: str

class WeightUpdateRequest(BaseModel):
    weight: float

# =======================================================
# 🌟 UPLOAD SAP FILE DENGAN RATE LIMITER
# =======================================================
@router.post("/orders/upload", response_model=schemas.UploadResponse)
@limiter.limit("5/hour")
async def upload_sap_file(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin_distribusi", "manager_logistik"))
):
    # Batas ukuran file upload (10 MB)
    MAX_UPLOAD_BYTES = 10 * 1024 * 1024
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File terlalu besar! Maksimal 10 MB, file kamu {len(contents) // (1024*1024)} MB."
        )
    settings = get_settings()
    
    try:
        count, success_list, failed_list = order_import_service.process_sap_file(contents, file.filename, db, settings)

        # Cek toko dari success_list yang tidak punya koordinat
        # success_list berisi list of dict — extract langsung, hindari N+1 query ke DB
        stores_without_coords = []
        for item in success_list:
            # Ambil dict item (bisa dict atau string tergantung versi)
            if not isinstance(item, dict):
                continue

            kordinat = str(item.get('kordinat') or item.get('koordinat') or '').strip()
            lat_val  = str(item.get('lat')      or '').strip()

            # Dianggap tidak punya koordinat kalau kordinat kosong, '-', atau 'None'
            has_coord = bool(
                (kordinat and kordinat not in ('', '-', 'None', 'none', '0,0')) or
                (lat_val  and lat_val  not in ('', '-', 'None', '0', '0.0'))
            )

            if not has_coord:
                # Fallback: cek MasterCustomer di DB
                kode = str(item.get('kode_customer') or '').strip()
                if kode:
                    cust = db.query(models.MasterCustomer).filter(
                        models.MasterCustomer.kode_customer == kode
                    ).first()
                    if cust and cust.latitude and cust.longitude:
                        has_coord = True  # MasterCustomer punya koordinat, skip

            if not has_coord:
                stores_without_coords.append({
                    "order_id":      item.get('order_id'),
                    "store_name":    item.get('nama_toko'),
                    "kode_customer": item.get('kode_customer'),
                })

        return {
            "message": (
                f"Upload selesai! {count} DO berhasil"
                + (f", {len(failed_list)} gagal" if failed_list else "")
                + (f". ⚠️ {len(stores_without_coords)} toko belum punya koordinat." if stores_without_coords else "")
                + "."
            ),
            "success_count": count,
            "success_list":  success_list,
            "failed_list":   failed_list,
            # Toko yang perlu di-pin sebelum bisa diikutkan VRP
            "stores_without_coordinates": stores_without_coords,
            "has_missing_coords": len(stores_without_coords) > 0,
        }
    except ValueError as ve:
        # Menangkap error validasi nilai data
        logger.error(f"⚠️ [UPLOAD SAP] Validation error oleh {current_user.username} di file {file.filename}: {str(ve)}")
        raise HTTPException(status_code=400, detail=str(ve))
    except pd.errors.EmptyDataError:
        # Menangkap error kalau file Excel/CSV kosong
        logger.error(f"⚠️ [UPLOAD SAP] File kosong diupload oleh {current_user.username} ({file.filename})")
        raise HTTPException(status_code=400, detail="File SAP yang diupload kosong atau formatnya tidak dikenali.")
    except Exception as e:
        # Fallback terakhir, tapi SEKARANG KITA LOG TRACEBACK-NYA FULL
        logger.error(f"🚨 [UPLOAD SAP FATAL] Gagal proses file {file.filename} oleh {current_user.username}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Terjadi kesalahan internal pada server saat memproses file.")

# =======================================================
# UPDATE OPERATIONS
# =======================================================
@router.put("/orders/{order_id}/time", response_model=schemas.OrderActionResponse)
def update_time_window(
    order_id: str, data: TimeUpdateRequest, db: Session = Depends(get_db), current_user: models.User = Depends(require_role("admin_distribusi", "manager_logistik"))
):
    try:
        order = OrderService.update_time_window(db, order_id, data.jam_maksimal)
        db.commit() 
        
        toko_name = order.customer.store_name if order.customer else "Toko"
        return {"message": f"Batas waktu {toko_name} diubah ke {data.jam_maksimal}", "order_id": order_id, "new_window_end": order.delivery_window_end}
    except OrderNotFoundError as e:
        db.rollback()
        raise HTTPException(status_code=404, detail=str(e))
    except OrderValidationError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/orders/{order_id}/coordinate", response_model=schemas.OrderActionResponse)
def update_coordinate(
    order_id: str, data: CoordinateUpdateRequest, db: Session = Depends(get_db), current_user: models.User = Depends(require_role("admin_distribusi", "manager_logistik"))
):
    try:
        OrderService.update_coordinate(db, order_id, data.latitude, data.longitude, data.kode_customer, data.nama_customer)
        db.commit() 
        
        if order_id.startswith("DRAFT-"):
            return {"message": "Koordinat DRAFT berhasil disimpan ke Master Database!", "order_id": order_id}
        return {"message": "Koordinat berhasil diupdate dan disimpan ke Master Database!", "order_id": order_id}
    except OrderNotFoundError as e:
        db.rollback()
        raise HTTPException(status_code=404, detail=str(e))

@router.put("/orders/{order_id}/weight", response_model=schemas.OrderActionResponse)
def update_weight(
    order_id: str,
    data: WeightUpdateRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin_distribusi", "manager_logistik")),
):
    try:
        OrderService.update_weight(db, order_id, data.weight)
        db.commit()
        return {"message": "Berat berhasil diupdate!", "order_id": order_id}
    except OrderNotFoundError as e:
        db.rollback()
        raise HTTPException(status_code=404, detail=str(e))
    except OrderValidationError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

# =======================================================
# APPROVE / REJECT POD DENGAN FSM VALIDATOR
# =======================================================
@router.put("/orders/{order_id}/pod/approve", response_model=schemas.PodVerificationResponse)
def approve_pod(
    order_id: str, data: schemas.PodApproveRequest, db: Session = Depends(get_db), current_user: models.User = Depends(require_role("admin_pod"))
):
    try:
        existing_order = db.query(models.DeliveryOrder).filter(
            models.DeliveryOrder.order_id == order_id
        ).first()

        if not existing_order:
            raise HTTPException(status_code=404, detail="DO tidak ditemukan.")

        validate_status_transition(
            existing_order.status,
            models.DOStatus.billed
        )

        order = OrderService.approve_pod(db, order_id)
        db.commit()
        return {"status": "success", "message": f"POD untuk DO {order_id} BERHASIL DISETUJUI!", "order_id": order_id, "new_status": order.status.value}
    except HTTPException:
        db.rollback()
        raise
    except OrderNotFoundError as e:
        db.rollback()
        raise HTTPException(status_code=404, detail=str(e))
    except OrderValidationError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/orders/{order_id}/pod/reject", response_model=schemas.PodVerificationResponse)
def reject_pod(
    order_id: str, data: schemas.PodRejectRequest, db: Session = Depends(get_db), current_user: models.User = Depends(require_role("admin_pod"))
):
    try:
        existing_order = db.query(models.DeliveryOrder).filter(
            models.DeliveryOrder.order_id == order_id
        ).first()

        if not existing_order:
            raise HTTPException(status_code=404, detail="DO tidak ditemukan.")

        validate_status_transition(
            existing_order.status,
            models.DOStatus.do_assigned_to_route
        )

        order = OrderService.reject_pod(db, order_id, data.reason, data.notes)
        db.commit()
        return {"status": "success", "message": f"POD untuk DO {order_id} DITOLAK! Alasan: {data.reason}", "order_id": order_id, "new_status": order.status.value}
    except HTTPException:
        db.rollback()
        raise
    except OrderNotFoundError as e:
        db.rollback()
        raise HTTPException(status_code=404, detail=str(e))
    except OrderValidationError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

# =======================================================
# READ-ONLY ENDPOINTS
# =======================================================
@router.post("/orders/upload-realisasi")
@limiter.limit("20/hour")
async def upload_realisasi(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin_distribusi", "manager_logistik"))
):
    """
    Upload Excel realisasi (jam ~1 malam) untuk update qty final dari gudang.
    Membaca kolom KODE CUST. + REALISASI dari Excel yang sama dengan routing.
    TIDAK re-route — hanya update weight_realisasi di delivery_orders.
    Fallback: jika tidak diupload, sistem tetap pakai weight_total (routing qty).
    """
    MAX_UPLOAD_BYTES = 10 * 1024 * 1024
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File terlalu besar! Maksimal 10 MB.")

    try:
        from io import BytesIO
        import re as _re

        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active

        # ── Deteksi header ────────────────────────────────────────────────
        def normalize(v):
            return _re.sub(r"[^A-Z0-9]", " ", str(v or "").upper()).strip()

        header_map = {}
        data_start_row = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                key = normalize(cell.value)
                if "KODE CUST" in key:
                    header_map["kode_customer"] = cell.column - 1
                    data_start_row = max(data_start_row or 0, cell.row + 1)
                if key in ("REALISASI", "QTY REALISASI", "REAL QTY"):
                    header_map["realisasi"] = cell.column - 1
                    data_start_row = max(data_start_row or 0, cell.row + 1)

        if "kode_customer" not in header_map:
            raise HTTPException(status_code=400,
                detail="Kolom 'KODE CUST.' tidak ditemukan. Pastikan format Excel sama dengan saat routing.")
        if "realisasi" not in header_map:
            raise HTTPException(status_code=400,
                detail="Kolom 'REALISASI' tidak ditemukan. Pastikan admin gudang sudah mengisi kolom REALISASI.")

        # ── Proses setiap baris ───────────────────────────────────────────
        import datetime as _dt
        today = _dt.date.today()

        updated, skipped_empty, not_found = [], [], []

        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            row_list = list(row)
            kode_raw  = row_list[header_map["kode_customer"]] if len(row_list) > header_map["kode_customer"] else None
            real_raw  = row_list[header_map["realisasi"]]      if len(row_list) > header_map["realisasi"]      else None

            if not kode_raw:
                continue

            kode = str(kode_raw).strip()

            # Skip baris dengan REALISASI kosong atau 0
            try:
                real_qty = float(str(real_raw).replace(",", ".")) if real_raw is not None else None
            except (ValueError, TypeError):
                real_qty = None

            if real_qty is None or real_qty == 0:
                skipped_empty.append(kode)
                continue

            # Cari DeliveryOrder untuk toko ini hari ini (routing sore hari)
            order = db.query(models.DeliveryOrder).join(
                models.MasterCustomer,
                models.DeliveryOrder.store_id == models.MasterCustomer.store_id,
                isouter=True
            ).filter(
                models.MasterCustomer.kode_customer == kode,
                models.DeliveryOrder.status.in_([
                    models.DOStatus.do_assigned_to_route,
                    models.DOStatus.delivered_pod_uploaded,
                    models.DOStatus.so_waiting_verification,
                    models.DOStatus.do_verified,
                ])
            ).order_by(models.DeliveryOrder.created_at.desc()).first()

            if not order:
                not_found.append(kode)
                continue

            # Update qty realisasi — HANYA ini, tidak menyentuh routing
            old_qty = order.weight_realisasi or order.weight_total
            order.weight_realisasi = real_qty
            updated.append({
                "kode_customer": kode,
                "order_id":      order.order_id,
                "qty_routing":   order.weight_total,
                "qty_realisasi": real_qty,
                "selisih_kg":    round(real_qty - (order.weight_total or 0), 1),
            })

        db.commit()

        total_selisih = sum(u["selisih_kg"] for u in updated)

        return {
            "status":        "success",
            "message":       f"✅ {len(updated)} order berhasil diupdate realisasinya.",
            "updated_count": len(updated),
            "updated":       updated,
            "skipped_empty": skipped_empty,
            "not_found":     not_found,
            "summary": {
                "total_selisih_kg": round(total_selisih, 1),
                "note": "Negatif = realisasi lebih kecil dari routing (normal). Positif = lebih besar.",
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"🚨 [UPLOAD REALISASI] Gagal: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500,
            detail=f"Gagal memproses file realisasi: {str(e)}")


@router.get("/orders", response_model=schemas.PendingOrderResponse)
def get_pending_orders(status: Optional[str] = None, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    query = db.query(models.DeliveryOrder).options(
        joinedload(models.DeliveryOrder.customer)
    )
    
    if status:
        try: query = query.filter(models.DeliveryOrder.status == models.DOStatus(status))
        except ValueError: pass

    orders = query.all()
    return {
        "status": "success", "total": len(orders),
        "data": [{
            "order_id": o.order_id, "customer_name": o.customer.store_name if o.customer else "Unknown", 
            "latitude": float(o.latitude) if o.latitude else None,
            "longitude": float(o.longitude) if o.longitude else None, "weight_total": o.weight_total,
            "delivery_window_start": o.delivery_window_start, "delivery_window_end": o.delivery_window_end,
            "status": o.status.value, "items": json.loads(o.service_type) if o.service_type and o.service_type.startswith('[') else []
        } for o in orders]
    }

@router.get("/pod/stats")
def get_pod_stats(db: Session = Depends(get_db),
                  current_user: models.User = Depends(require_role("admin_pod", "manager_logistik"))):
    """
    KPI stats untuk POD dashboard:
    - pending: DO masih menunggu review admin POD
    - auto_verified_pct: % DO yang langsung success tanpa rejection
    - rejected_today: DO yang failed/rejected hari ini
    - avg_processing_mins: rata-rata waktu antrean (menit sejak pod_uploaded)
    """
    import datetime as _dt
    today = _dt.date.today()

    # Pending — menunggu verifikasi admin POD
    pending = db.query(models.DeliveryOrder).filter(
        models.DeliveryOrder.status == models.DOStatus.delivered_pod_uploaded
    ).count()

    # Total selesai hari ini (success + partial + failed)
    total_completed = db.query(models.DeliveryOrder).join(
        models.TMSRouteLine, models.TMSRouteLine.order_id == models.DeliveryOrder.order_id
    ).join(
        models.TMSRoutePlan, models.TMSRoutePlan.route_id == models.TMSRouteLine.route_id
    ).filter(
        models.TMSRoutePlan.planning_date == today,
        models.DeliveryOrder.status.in_([
            models.DOStatus.delivered_success,
            models.DOStatus.delivered_partial,
            models.DOStatus.failed,
        ])
    ).count()

    # Success (auto-verified = tidak perlu manual rejection)
    total_success = db.query(models.DeliveryOrder).join(
        models.TMSRouteLine, models.TMSRouteLine.order_id == models.DeliveryOrder.order_id
    ).join(
        models.TMSRoutePlan, models.TMSRoutePlan.route_id == models.TMSRouteLine.route_id
    ).filter(
        models.TMSRoutePlan.planning_date == today,
        models.DeliveryOrder.status.in_([
            models.DOStatus.delivered_success,
            models.DOStatus.delivered_partial,
        ])
    ).count()

    # Rejected hari ini
    rejected_today = db.query(models.DeliveryOrder).join(
        models.TMSRouteLine, models.TMSRouteLine.order_id == models.DeliveryOrder.order_id
    ).join(
        models.TMSRoutePlan, models.TMSRoutePlan.route_id == models.TMSRouteLine.route_id
    ).filter(
        models.TMSRoutePlan.planning_date == today,
        models.DeliveryOrder.status == models.DOStatus.failed
    ).count()

    # Auto-verified %
    auto_verified_pct = round((total_success / total_completed * 100), 1) if total_completed > 0 else 0.0

    # Avg processing: rata-rata usia antrean yang pending (menit sejak created_at sebagai proxy)
    pending_orders = db.query(models.DeliveryOrder).filter(
        models.DeliveryOrder.status == models.DOStatus.delivered_pod_uploaded
    ).all()
    if pending_orders:
        now = _dt.datetime.now()
        ages = []
        for o in pending_orders:
            if o.created_at:
                age_min = (now - o.created_at).total_seconds() / 60
                ages.append(age_min)
        avg_processing = round(sum(ages) / len(ages), 1) if ages else 0.0
    else:
        avg_processing = 0.0

    return {
        "status": "success",
        "data": {
            "pending":            pending,
            "auto_verified_pct":  auto_verified_pct,
            "rejected_today":     rejected_today,
            "total_completed":    total_completed,
            "avg_queue_mins":     avg_processing,  # menit rata-rata dalam antrean
        }
    }


@router.get("/pod/verifications")
def get_pod_verifications(db: Session = Depends(get_db), current_user: models.User = Depends(require_role("admin_pod"))):
    orders = db.query(models.DeliveryOrder).options(
        joinedload(models.DeliveryOrder.customer),
        joinedload(models.DeliveryOrder.route_line).joinedload(models.TMSRouteLine.epod),
        joinedload(models.DeliveryOrder.route_line).joinedload(models.TMSRouteLine.route_plan).joinedload(models.TMSRoutePlan.driver),
        joinedload(models.DeliveryOrder.route_line).joinedload(models.TMSRouteLine.route_plan).joinedload(models.TMSRoutePlan.vehicle)
    ).filter(
        models.DeliveryOrder.status == models.DOStatus.delivered_pod_uploaded
    ).all()
    
    data = []
    for o in orders:
        customer = o.customer
        route_line = o.route_line
        epod = route_line.epod if route_line else None
        route_plan = route_line.route_plan if route_line else None
        driver = route_plan.driver if route_plan else None
        vehicle = route_plan.vehicle if route_plan else None
        
        items = json.loads(o.service_type) if o.service_type and o.service_type.startswith('[') else []
        
        data.append({
            "order_id": o.order_id, "customer_name": customer.store_name if customer else "Unknown Store",
            "customer_address": customer.address if customer else "Unknown Address", "driver_name": driver.name if driver else "Unknown Driver",
            "driver_phone": driver.phone if driver else "", "vehicle_plate": vehicle.license_plate if vehicle else "",
            "vehicle_type": vehicle.type if vehicle else "", "photo_url": epod.photo_url if epod else "",
            "gps_lat": float(epod.gps_location_lat) if (epod and epod.gps_location_lat) else None,
            "gps_lon": float(epod.gps_location_lon) if (epod and epod.gps_location_lon) else None,
            "qty_delivered": epod.qty_delivered if epod else (o.weight_total or 0.0),
            "qty_return": epod.qty_return if epod else 0.0, "qty_damaged": epod.qty_damaged if epod else 0.0,
            "return_reason": epod.return_reason if epod else "", "driver_notes": epod.driver_notes if epod else "",
            "timestamp": epod.timestamp.strftime("%Y-%m-%d %H:%M:%S") if (epod and epod.timestamp) else "", "items": items
        })
    return {"status": "success", "total": len(data), "data": data}

@router.get("/pod/history")
def get_pod_history(status: Optional[str] = None, search: Optional[str] = None, db: Session = Depends(get_db), current_user: models.User = Depends(require_role("admin_pod"))):
    query = db.query(models.DeliveryOrder).options(
        joinedload(models.DeliveryOrder.customer),
        joinedload(models.DeliveryOrder.route_line).joinedload(models.TMSRouteLine.epod),
        joinedload(models.DeliveryOrder.route_line).joinedload(models.TMSRouteLine.route_plan).joinedload(models.TMSRoutePlan.driver),
        joinedload(models.DeliveryOrder.route_line).joinedload(models.TMSRouteLine.route_plan).joinedload(models.TMSRoutePlan.vehicle)
    ).filter(
        models.DeliveryOrder.status.in_([models.DOStatus.billed, models.DOStatus.delivered_success, models.DOStatus.delivered_partial])
    )
    
    if status and status != "ALL":
        try:
            if status == "SUCCESS": query = query.filter(models.DeliveryOrder.status == models.DOStatus.billed)
            elif status == "PARTIAL": query = query.filter(models.DeliveryOrder.status == models.DOStatus.delivered_partial)
            elif status == "DELIVERED_SUCCESS": query = query.filter(models.DeliveryOrder.status == models.DOStatus.delivered_success)
        except Exception: pass
            
    orders = query.all()
    
    data = []
    for o in orders:
        customer = o.customer
        route_line = o.route_line
        epod = route_line.epod if route_line else None
        route_plan = route_line.route_plan if route_line else None
        driver = route_plan.driver if route_plan else None
        vehicle = route_plan.vehicle if route_plan else None
        
        driver_name = driver.name if driver else "Unknown Driver"
        customer_name = customer.store_name if customer else "Unknown Store"
        
        if search:
            s = search.lower()
            if s not in o.order_id.lower() and s not in driver_name.lower() and s not in customer_name.lower(): continue
                
        items = json.loads(o.service_type) if o.service_type and o.service_type.startswith('[') else []
                
        data.append({
            "order_id": o.order_id, "customer_name": customer_name, "customer_address": customer.address if customer else "Unknown Address",
            "driver_name": driver_name, "driver_phone": driver.phone if driver else "", "vehicle_plate": vehicle.license_plate if vehicle else "",
            "vehicle_type": vehicle.type if vehicle else "", "photo_url": epod.photo_url if epod else "",
            "gps_lat": float(epod.gps_location_lat) if (epod and epod.gps_location_lat) else None,
            "gps_lon": float(epod.gps_location_lon) if (epod and epod.gps_location_lon) else None,
            "qty_delivered": epod.qty_delivered if epod else (o.weight_total or 0.0),
            "qty_return": epod.qty_return if epod else 0.0, "qty_damaged": epod.qty_damaged if epod else 0.0,
            "return_reason": epod.return_reason if epod else "", "driver_notes": epod.driver_notes if epod else "",
            "status": o.status.value, "timestamp": epod.timestamp.strftime("%Y-%m-%d %H:%M:%S") if (epod and epod.timestamp) else "",
            "items": items
        })
    return {"status": "success", "total": len(data), "data": data}