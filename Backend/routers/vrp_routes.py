# Backend/routers/vrp_routes.py
import datetime
import json
import re
import os
import shutil
import logging
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy.exc import SQLAlchemyError

import models
import schemas
from services import osrm_service, eta_service
from services.job_store import VRP_JOBS, TRAFFIC_JOBS
from dependencies import get_db, get_settings, get_current_user, require_role


logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api", tags=["Route Management & Export"])

# ==========================================
# 1. GET ROUTES
# ==========================================
@router.get("/routes", response_model=schemas.GetRoutesResponse)
def get_routes(
    date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    # 🌟 FIX N+1: Tambahkan eager loading untuk memuat semua relasi di awal
    query = db.query(models.TMSRoutePlan).options(
        joinedload(models.TMSRoutePlan.driver),
        joinedload(models.TMSRoutePlan.vehicle),
        # Asumsi relasi di model TMSRoutePlan ke TMSRouteLine bernama 'route_lines'
        # Gunakan selectinload untuk One-to-Many (koleksi)
        selectinload(models.TMSRoutePlan.route_lines)
        .joinedload(models.TMSRouteLine.order)
        .joinedload(models.DeliveryOrder.customer)
    )
    
    if date:
        try:
            query = query.filter(
                models.TMSRoutePlan.planning_date
                == datetime.datetime.strptime(date, "%Y-%m-%d").date()
            )
        except ValueError:
            logger.warning(f"⚠️ Invalid route date format: {date}")
            raise HTTPException(status_code=400, detail="Format tanggal salah! Gunakan YYYY-MM-DD.")

    zona_dummy = [
        "Bekasi/Cikarang", "Kelapa Gading", "Kembangan/Jakbar",
        "Serpong/BSD", "Pusat/Selatan", "Bogor", "Tigaraksa",
    ]
    settings = get_settings()
    routes = query.all()
    is_route_empty = len(routes) == 0
    hasil = []

    for rute in routes:
        detail_rute = []
        
        # 🌟 FIX N+1: Ambil lines dari relasi, BUKAN query db lagi di dalam loop
        # Jangan lupa di-sort by sequence menggunakan python jika belum tersort dari relasi
        lines = sorted(rute.route_lines, key=lambda x: x.sequence) if rute.route_lines else []
        
        for line in lines:
            order = line.order # 🌟 FIX N+1: Akses via relasi
            if not order:
                continue

            items = (
                json.loads(order.service_type)
                if order.service_type and order.service_type.startswith("[")
                else []
            )
            nama_toko = order.customer.store_name if order.customer else "Toko"

            berat = float(order.weight_total) if order.weight_total else 0.0
            detail_rute.append({
                "urutan": line.sequence,
                "nama_toko": nama_toko,
                "latitude": float(order.latitude) if order.latitude else 0.0,
                "longitude": float(order.longitude) if order.longitude else 0.0,
                "lat": float(order.latitude) if order.latitude else 0.0,
                "lon": float(order.longitude) if order.longitude else 0.0,
                "berat_kg": berat,
                "turun_barang_kg": berat,   # alias — dibutuhkan schema & frontend
                "jam_tiba": str(line.est_arrival),
                "distance_from_prev_km": float(line.distance_from_prev_km or 0.0),
                "items": items,
                "nomor_do": order.order_id,
                "order_id": order.order_id,
            })

        # Baca geometry dari kolom DB (bukan file yang tidak pernah ditulis)
        garis_aspal = []
        try:
            if rute.route_geometry:
                garis_aspal = json.loads(rute.route_geometry)
        except Exception:
            pass

        # Ekstrak nomor truk dari format "RP-YYYYMMDD-T{n}" dengan regex
        # Fallback ke hash untuk format route_id lain (distribusi merata)
        _zona_match = re.search(r'T(\d+)$', rute.route_id)
        if _zona_match:
            idx_zona = (int(_zona_match.group(1)) - 1) % len(zona_dummy)
        else:
            idx_zona = abs(hash(rute.route_id)) % len(zona_dummy)
        transport_cost = round(
            (rute.total_distance_km or 0)
            * (settings.cost_fuel_per_liter / settings.cost_avg_km_per_liter)
        )

        hasil.append({
            "route_id": rute.route_id,
            "id": rute.route_id,
            "tanggal": str(rute.planning_date),
            "driver_name": rute.driver.name if rute.driver else "-",
            "kendaraan": rute.vehicle.license_plate if rute.vehicle else "-",
            "armada": rute.vehicle.license_plate if rute.vehicle else "-",
            "vehicle": rute.vehicle.license_plate if rute.vehicle else "-",
            "jenis": rute.vehicle.type if rute.vehicle else "-",
            "destinasi_jumlah": len(detail_rute),
            "total_berat": rute.total_weight,
            "total_muatan_kg": rute.total_weight or 0.0,
            "total_weight": rute.total_weight or 0.0,
            "total_distance_km": rute.total_distance_km,
            "total_jarak_km": rute.total_distance_km or 0.0,
            "transport_cost": transport_cost,
            "status": "Aktif",
            "zone": zona_dummy[idx_zona],
            "detail_rute": detail_rute,
            "detail_perjalanan": detail_rute,
            "garis_aspal": garis_aspal,
        })

    dropped_nodes_response = []
    if not is_route_empty:
        # 🌟 FIX N+1: Joinedload customer untuk DO yang belum ter-assign
        unassigned = (
            db.query(models.DeliveryOrder)
            .options(joinedload(models.DeliveryOrder.customer))
            .filter(models.DeliveryOrder.status == models.DOStatus.do_verified)
            .all()
        )
        dropped_nodes_response = [
            {
                "nama_toko": o.customer.store_name if o.customer else "Toko",
                "berat_kg": o.weight_total,
                "alasan": "Drop AI (Overcapacity)",
                "lat": float(o.latitude) if o.latitude else 0.0,
                "lon": float(o.longitude) if o.longitude else 0.0,
            }
            for o in unassigned
        ]

    return {"routes": hasil, "dropped_nodes": dropped_nodes_response}


# ==========================================
# 2. RESEQUENCE (TSP) SETELAH MANUAL OVERRIDE
# ==========================================
@router.post("/routes/resequence")
def resequence_routes(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin_distribusi", "manager_logistik")),
):
    from utils.helpers import time_str_to_minutes, menit_ke_jam
    from services import vrp_solver

    try:
        settings = get_settings()
        DEPO_LAT = settings.depo_lat
        DEPO_LON = settings.depo_lon
        START_MINUTE = time_str_to_minutes(settings.vrp_start_time)
        BASE_DROP = settings.vrp_base_drop_time_mins
        VAR_DROP = settings.vrp_var_drop_time_mins

        jadwal = payload.get("jadwal_truk_internal", [])

        for truk in jadwal:
            stops = truk.get("detail_perjalanan", [])
            customers = [
                s for s in stops
                if str(s.get("keterangan", "")).lower() not in ["start", "finish"]
                # Oncall stops tidak punya "urutan" eksplisit — filter by lat/lon valid
                and s.get("lat") is not None
                and s.get("lon") is not None
                and not str(s.get("nama_toko", "") or s.get("lokasi", "")).upper().startswith("GUDANG")
            ]

            if len(customers) < 2:
                continue

            locations = [{"lat": DEPO_LAT, "lon": DEPO_LON}]
            demands = [0]
            customer_map = {}

            for idx, c in enumerate(customers):
                locations.append({"lat": float(c["lat"]), "lon": float(c["lon"])})
                berat = float(c.get("turun_barang_kg", 0) or c.get("berat_kg", 0))
                demands.append(int(berat))
                customer_map[idx + 1] = c

            # Pass departure_hour aktual agar traffic factor sesuai jam resequence
            _reseq_hour = datetime.now().hour
            dist_mat, time_mat = osrm_service.build_osrm_matrix(locations, departure_hour=_reseq_hour)
            if not dist_mat:
                dist_mat, time_mat = osrm_service.build_haversine_matrix(locations, departure_hour=_reseq_hour)

            # Pakai dist_mat langsung (meter) — tidak ada lagi pembulatan ke km
            # yang membuang presisi untuk toko-toko berdekatan (<1 km)
            caps = [sum(demands) + 9999]
            is_mall = [False] * len(locations)
            time_windows = [(0, 1440)] * len(locations)

            hasil_tsp = vrp_solver.solve_vrp(
                dist_mat, time_mat, demands, 1, caps, is_mall, time_windows, BASE_DROP, VAR_DROP
            )

            best_indices = (
                hasil_tsp["routes"][0]
                if hasil_tsp and hasil_tsp["routes"] and len(hasil_tsp["routes"][0]) > 0
                else list(range(len(locations)))
            )

            new_manifest = []
            current_time = START_MINUTE
            prev_node = 0
            total_jarak_m = 0

            for step, node_idx in enumerate(best_indices):
                seg_km = round((dist_mat[prev_node][node_idx] if step != 0 else 0) / 1000.0, 1)
                if step != 0:
                    current_time += time_mat[prev_node][node_idx]
                    total_jarak_m += dist_mat[prev_node][node_idx]

                if node_idx == 0:
                    new_manifest.append({
                        "urutan": step,
                        "lokasi": "📍 GUDANG JAPFA",
                        "jam": str(menit_ke_jam(current_time)),
                        "keterangan": "Start" if step == 0 else "Finish",
                        "lat": DEPO_LAT,
                        "lon": DEPO_LON,
                        "distance_from_prev_km": seg_km,
                    })
                else:
                    cust = customer_map[node_idx]
                    cust["urutan"] = step
                    cust["jam_tiba"] = str(menit_ke_jam(current_time))
                    cust["distance_from_prev_km"] = seg_km
                    new_manifest.append(cust)

                    service_time = BASE_DROP + (demands[node_idx] * VAR_DROP / 10.0)
                    current_time += service_time

                prev_node = node_idx

            # Tambah entry Finish ke depot
            finish_m = dist_mat[prev_node][0] if dist_mat else 0
            current_time += time_mat[prev_node][0] if time_mat else 0
            total_jarak_m += finish_m

            new_manifest.append({
                "urutan": len(best_indices),
                "lokasi": "📍 GUDANG JAPFA",
                "jam": str(menit_ke_jam(current_time)),
                "keterangan": "Finish",
                "lat": DEPO_LAT,
                "lon": DEPO_LON,
                "distance_from_prev_km": round(finish_m / 1000.0, 1),
            })

            route_geometry = osrm_service.get_road_geometry(best_indices + [0], locations)
            truk["garis_aspal"] = route_geometry
            truk["detail_perjalanan"] = new_manifest
            truk["total_jarak_km"] = round(total_jarak_m / 1000.0, 1)

        return payload

    except Exception as e:
        logger.error(f"🚨 [RESEQUENCE ERROR]: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Gagal menghitung ulang urutan rute.")


# ==========================================
# 3. CONFIRM ROUTES — SIMPAN KE DATABASE
# ==========================================
@router.post("/routes/confirm", response_model=schemas.ConfirmRouteResponse)
def confirm_routes(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin_distribusi", "manager_logistik")),
):

    today = datetime.datetime.now().date()
    jadwal = payload.get("jadwal_truk_internal", [])

    if not jadwal:
        raise HTTPException(status_code=400, detail="Tidak ada data rute untuk dikonfirmasi.")

    # =====================================================
    # 🌟 PHASE 1 — VALIDASI SEMUA (MEMBUKA TRANSAKSI SECARA OTOMATIS)
    # =====================================================
    used_route_ids: set = set()

    for truk in jadwal:
        route_id = truk.get("route_id")
        nopol = truk.get("armada")
        drv_id = truk.get("driver_id")
        hlp_id = truk.get("helper_id")

        if not route_id:
            raise HTTPException(status_code=400, detail="Ada route tanpa route_id.")
        if route_id in used_route_ids:
            raise HTTPException(status_code=400, detail=f"Duplicate route_id: {route_id}")
        used_route_ids.add(route_id)

        vehicle = db.query(models.FleetVehicle).filter(models.FleetVehicle.license_plate == nopol).first()
        if not vehicle:
            raise HTTPException(status_code=400, detail=f"Armada '{nopol}' tidak ditemukan.")

        if not drv_id:
            raise HTTPException(status_code=400, detail=f"Truk {nopol} belum diassign driver.")

        driver = db.query(models.HRDriver).filter(models.HRDriver.driver_id == drv_id).first()
        if not driver:
            raise HTTPException(status_code=400, detail=f"Driver ID {drv_id} tidak ditemukan.")

        if str(hlp_id).lower() in ["103", "0", "", "none", "null"]:
            hlp_id = None
            
        if hlp_id:
            helper = db.query(models.HRDriver).filter(models.HRDriver.driver_id == hlp_id).first()
            if not helper:
                raise HTTPException(status_code=400, detail=f"Helper ID {hlp_id} tidak ditemukan.")

        for stop in truk.get("detail_perjalanan", []):
            nomor_do = stop.get("nomor_do") or stop.get("order_id") or stop.get("id")
            if not nomor_do:
                continue
            order = db.query(models.DeliveryOrder).filter(models.DeliveryOrder.order_id == nomor_do).first()
            if not order:
                raise HTTPException(status_code=400, detail=f"DO '{nomor_do}' tidak ditemukan.")

    # =====================================================
    # 🌟 PHASE 2 — EKSEKUSI DATABASE
    # (HAPUS `with db.begin():`, gunakan db.commit() dan db.rollback())
    # =====================================================
    try:
        # 1. AMBIL SEMUA RUTE LAMA HARI INI
        existing_routes = db.query(models.TMSRoutePlan).filter(
            models.TMSRoutePlan.planning_date == today
        ).all()

        existing_route_ids = [r.route_id for r in existing_routes]

        # 2. HAPUS ROUTE LINE DULU (BULK DELETE)
        if existing_route_ids:
            db.query(models.TMSRouteLine).filter(
                models.TMSRouteLine.route_id.in_(existing_route_ids)
            ).delete(synchronize_session=False)

        # 3. HAPUS ROUTE PLAN
        db.query(models.TMSRoutePlan).filter(
            models.TMSRoutePlan.planning_date == today
        ).delete(synchronize_session=False)

        # 4. INSERT RUTE BARU
        for truk in jadwal:
            route_id = truk.get("route_id")
            nopol = truk.get("armada")

            existing = db.query(models.TMSRoutePlan).filter(
                models.TMSRoutePlan.route_id == route_id
            ).first()

            if existing:
                raise ValueError(f"Duplicate route_id detected: {route_id}")

            vehicle = db.query(models.FleetVehicle).filter(
                models.FleetVehicle.license_plate == nopol
            ).first()

            if not vehicle:
                raise ValueError(f"Armada '{nopol}' tidak ditemukan di master kendaraan!")

            hlp_id = truk.get("helper_id")
            if str(hlp_id).lower() in ["103", "0", "", "none", "null"]:
                hlp_id = None

            # Simpan geometry ke DB (menggantikan file route_geometries/*.json)
            _garis = truk.get("garis_aspal") or []
            _geo_json = json.dumps(_garis) if _garis else None

            new_plan = models.TMSRoutePlan(
                route_id=route_id,
                planning_date=today,
                vehicle_id=vehicle.vehicle_id,
                driver_id=truk.get("driver_id"),
                helper_id=hlp_id,
                total_weight=truk.get("total_muatan_kg", 0),
                total_distance_km=truk.get("total_jarak_km", 0),
                route_geometry=_geo_json,
            )
            db.add(new_plan)

            for stop in truk.get("detail_perjalanan", []):
                nomor_do = stop.get("nomor_do") or stop.get("order_id") or stop.get("id")
                if not nomor_do:
                    continue

                try:
                    jam_parts = str(stop["jam_tiba"]).split(":")
                    jam_est = datetime.time(hour=int(jam_parts[0]), minute=int(jam_parts[1]))
                except Exception as e:
                    logger.warning(
                        f"⚠️ Invalid jam_tiba pada route {route_id}: {stop.get('jam_tiba')} | {str(e)}"
                    )
                    jam_est = datetime.time(hour=12, minute=0)

                route_line = models.TMSRouteLine(
                    route_id=route_id,
                    order_id=nomor_do,
                    sequence=stop.get("urutan", 0),
                    est_arrival=jam_est,
                    distance_from_prev_km=stop.get("distance_from_prev_km", 0),
                )
                db.add(route_line)

                order = db.query(models.DeliveryOrder).filter(
                    models.DeliveryOrder.order_id == nomor_do
                ).first()

                if not order:
                    raise ValueError(f"Delivery Order '{nomor_do}' tidak ditemukan!")

                order.status = models.DOStatus.do_assigned_to_route

        # 🌟 SEMUA PROSES BERHASIL, SAATNYA COMMIT!
        db.commit()
        logger.info(f"✅ [CONFIRM] {len(jadwal)} rute dikonfirmasi untuk {today} oleh {current_user.username}")

        return {
            "message": f"Sukses! {len(jadwal)} rute berhasil dikonfirmasi.",
            "status": "success"
        }

    # =====================================================
    # 🌟 DATABASE ERROR (ROLLBACK)
    # =====================================================
    except SQLAlchemyError as db_err:
        db.rollback() # 🌟 Tarik mundur semua perubahan!
        logger.error(
            f"🚨 [DB ERROR] Confirm routes gagal oleh {current_user.username}. Error: {str(db_err)}",
            exc_info=True
        )
        raise HTTPException(
            status_code=500,
            detail="Gagal menyimpan rute ke database. Perubahan telah di-rollback."
        )

    # =====================================================
    # 🌟 GENERAL ERROR (ROLLBACK)
    # =====================================================
    except Exception as e:
        db.rollback() # 🌟 Tarik mundur semua perubahan!
        logger.error(
            f"🚨 [CONFIRM ROUTES FATAL] Error sistem saat {current_user.username} confirm {len(jadwal)} rute. Error: {str(e)}",
            exc_info=True
        )
        raise HTTPException(
            status_code=500,
            detail="Gagal mengonfirmasi rute. Sistem telah otomatis me-rollback perubahan untuk mencegah korupsi data."
        )
# ==========================================
# 4. EXPORT KE CORPORATE EXCEL TEMPLATE
# ==========================================
@router.get("/routes/export-excel", response_class=FileResponse)
def export_all_routes_to_excel(
    date: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin_distribusi", "manager_logistik")),
):
    try:
        target_date = datetime.datetime.strptime(date, "%Y-%m-%d").date()
        routes = db.query(models.TMSRoutePlan).filter(
            models.TMSRoutePlan.planning_date == target_date
        ).all()

        if not routes:
            raise HTTPException(status_code=400, detail="Tidak ada rute berjalan di tanggal ini!")

        template_path = "templates/DO_TEMPLATE.xlsx"
        if not os.path.exists(template_path):
            raise HTTPException(
                status_code=404,
                detail="File DO_TEMPLATE.xlsx tidak ditemukan di folder Backend/templates/"
            )

        os.makedirs("static/uploads", exist_ok=True)
        output_filename = f"Surat_Jalan_JAPFA_{date}.xlsx"
        output_path = f"static/uploads/{output_filename}"
        shutil.copyfile(template_path, output_path)

        wb = openpyxl.load_workbook(output_path)
        sheet_name = "JADWAL" if "JADWAL" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        def tulis_sel(col_letter: str, row_number: int, val):
            coord = f"{col_letter}{row_number}"
            cell = ws[coord]
            if type(cell).__name__ == "MergedCell":
                for m_range in ws.merged_cells.ranges:
                    if coord in m_range:
                        ws.cell(row=m_range.min_row, column=m_range.min_col).value = val
                        break
            else:
                cell.value = val

        BLOCK_HEIGHT = 36

        for idx, rute in enumerate(routes):
            offset = idx * BLOCK_HEIGHT

            if idx == 0:
                tulis_sel("I", 2, str(target_date))

            tulis_sel("C", 7 + offset, rute.vehicle.license_plate if rute.vehicle else "-")
            tulis_sel("J", 7 + offset, rute.driver.name if rute.driver else "-")
            tulis_sel("J", 8 + offset, rute.helper.name if rute.helper else "Tanpa Helper")

            lines = (
                db.query(models.TMSRouteLine)
                .filter(models.TMSRouteLine.route_id == rute.route_id)
                .order_by(models.TMSRouteLine.sequence)
                .all()
            )

            for item_idx, line in enumerate(lines):
                if item_idx >= 30:
                    break

                row_target = 10 + offset + item_idx
                order = line.order
                if not order:
                    continue

                nama_toko = "Toko JAPFA"
                kode_toko = ""
                if order.customer:
                    nama_toko = order.customer.store_name
                    kode_toko = order.customer.kode_customer or ""

                jenis_barang = "FROZEN"
                if order.service_type and order.service_type.startswith("["):
                    try:
                        items = json.loads(order.service_type)
                        if items:
                            jenis_barang = items[0].get("tipe", "FROZEN")
                    except Exception:
                        pass

                tulis_sel("A", row_target, rute.vehicle.license_plate if rute.vehicle else "-")
                tulis_sel("B", row_target, line.sequence)
                tulis_sel("C", row_target, kode_toko)
                tulis_sel("D", row_target, nama_toko)
                tulis_sel("E", row_target, float(order.weight_total) if order.weight_total else 0)
                tulis_sel("F", row_target, jenis_barang)
                tulis_sel("G", row_target, "AYAM")
                tulis_sel("H", row_target, str(line.est_arrival)[:5] if line.est_arrival else "-")
                tulis_sel("K", row_target, order.order_id)

            tulis_sel("E", 41 + offset, rute.total_weight)

        wb.save(output_path)

        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"🚨 [EXCEL EXPORT ERROR]: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Gagal generate Excel: {str(e)}")