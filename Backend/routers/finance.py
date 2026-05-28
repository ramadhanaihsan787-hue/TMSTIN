# Backend/routers/finance.py
import io
import uuid
import json
import logging
import calendar
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import models
import schemas
from dependencies import get_db, get_current_user, require_role

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/finance", tags=["Finance & Expenses"])

# ==========================================
# HELPER FUNCTIONS
# ==========================================
def expense_to_dict(e: models.OperationalExpense) -> dict:
    """Helper buat mapping data DB ke format Frontend biar DRY dengan JSON Serialization Fallback"""
    plate = "N/A"
    vehicle_type = "N/A"
    driver = "N/A"
    helper = e.helper_name or ""
    notes = e.notes or ""

    if e.notes:
        try:
            # Coba parse notes sebagai JSON block untuk temporary/on-call data
            notes_data = json.loads(e.notes)
            if isinstance(notes_data, dict):
                plate = notes_data.get("plate", "N/A")
                vehicle_type = notes_data.get("vehicleType", "N/A")
                driver = notes_data.get("driver", "N/A")
                helper = notes_data.get("helper", helper)
                notes = notes_data.get("notes", "")
        except Exception as err:
            logger.warning(f"⚠️ Gagal parse notes JSON untuk expense {e.id}: {str(err)}")

    # Fallback ke real DB relationships jika plate/driver masih default N/A
    if plate == "N/A" and e.vehicle:
        plate = e.vehicle.license_plate or "N/A"
        vehicle_type = getattr(e.vehicle, "type", "CDD") or "CDD"
    if driver == "N/A" and e.driver:
        driver = getattr(e.driver, "name", "N/A") or "N/A"

    return {
        "id": e.id,
        "time": e.time,
        "date": str(e.date),
        "plate": plate,
        "vehicleType": vehicle_type,
        "driver": driver,
        "isOncall": e.is_oncall,
        "bbm": e.bbm,
        "tol": e.tol,
        "parkir": e.parkir,
        "parkirLiar": e.parkir_liar,
        "kuliAngkut": e.kuli_angkut,
        "lainLain": e.lain_lain,
        "helperName": helper,
        "notes": notes,
        "total": e.total,
        "kmAwal": e.km_awal,
        "kmAkhir": e.km_akhir,
        "jamBerangkat": e.jam_berangkat,
        "jamPulang": e.jam_pulang,
    }

# ==========================================
# MASTER DATA ENDPOINTS
# ==========================================
@router.get("/master-data")
def get_finance_master_data(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Retrieve actual active vehicles and drivers list for cashier drop-downs"""
    vehicles = db.query(models.FleetVehicle).all()
    drivers = db.query(models.HRDriver).all()
    return {
        "status": "success",
        "data": {
            "fleets": [
                {
                    "id": v.vehicle_id,
                    "plate": v.license_plate,
                    "type": v.type or "CDD"
                } for v in vehicles
            ],
            "drivers": [
                {
                    "id": d.driver_id,
                    "name": d.name
                } for d in drivers
            ]
        }
    }

# 0. BOP AUTOFILL
@router.get("/bop-autofill")
def get_bop_autofill(
    plate: str,
    tanggal: str = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Auto-fill data perjalanan dari TMSRoutePlan hari ini untuk plat nomor tertentu."""
    target_date = date.today()
    if tanggal:
        try:
            target_date = datetime.strptime(tanggal, "%Y-%m-%d").date()
        except ValueError:
            pass

    vehicle = db.query(models.FleetVehicle).filter(models.FleetVehicle.license_plate == plate).first()
    if not vehicle:
        return {"status": "success", "data": None, "message": "Kendaraan tidak ditemukan di master armada"}

    plan = db.query(models.TMSRoutePlan).filter(
        models.TMSRoutePlan.vehicle_id == vehicle.vehicle_id,
        models.TMSRoutePlan.planning_date == target_date
    ).first()

    if not plan:
        return {"status": "success", "data": None, "message": "Belum ada rute terdispatch untuk armada ini hari ini"}

    driver_name = plan.driver.name if plan.driver else None
    helper_name = plan.helper.name if plan.helper else None

    jam_berangkat = None
    if plan.start_time:
        default_06 = plan.start_time.replace(hour=6, minute=0, second=0, microsecond=0)
        if plan.start_time != default_06:
            jam_berangkat = plan.start_time.strftime("%H:%M")

    jam_pulang = plan.end_time.strftime("%H:%M") if plan.end_time else None

    return {
        "status": "success",
        "data": {
            "plate": vehicle.license_plate,
            "vehicle_type": vehicle.type or "CDD",
            "vehicle_id": vehicle.vehicle_id,
            "driver_name": driver_name,
            "driver_id": plan.driver_id,
            "helper_name": helper_name,
            "helper_id": plan.helper_id,
            "route_id": plan.route_id,
            "tanggal": str(target_date),
            "jam_berangkat": jam_berangkat,
            "jam_pulang": jam_pulang,
            "km_awal": plan.km_awal_trip,
            "km_akhir": plan.km_akhir_trip,
            "source": "driver_app" if (jam_berangkat or plan.km_awal_trip) else "belum_ada_data",
        }
    }


# ==========================================
# BOP EXPORT & IMPORT
# ==========================================
@router.get("/bop-export")
def export_bop(
    month: int = None,
    year: int = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Generate file Excel BOP format resmi JAPFA WH Cikupa."""
    today = datetime.now()
    m = month or today.month
    y = year or today.year

    from_date = date(y, m, 1)
    to_date = date(y, m, calendar.monthrange(y, m)[1])

    expenses = db.query(models.OperationalExpense).filter(
        models.OperationalExpense.date >= from_date,
        models.OperationalExpense.date <= to_date
    ).order_by(models.OperationalExpense.date).all()

    BULAN = ['', 'JANUARI','FEBRUARI','MARET','APRIL','MEI','JUNI',
             'JULI','AGUSTUS','SEPTEMBER','OKTOBER','NOVEMBER','DESEMBER']

    wb = Workbook()
    ws = wb.active
    ws.title = f"{m:02d}"

    # Style Helpers
    def thin_border():
        t = Side(style='thin')
        return Border(left=t, right=t, top=t, bottom=t)
        
    def cell(row, col, val="", bold=False, center=True, bg=None, font_size=9):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name='Calibri', bold=bold, size=font_size)
        c.alignment = Alignment(
            horizontal='center' if center else 'left', 
            vertical='center', 
            wrap_text=True
        )
        c.border = thin_border()
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        return c

    # Row 1: Judul
    ws.merge_cells('A1:T1')
    t = ws['A1']
    t.value = "REKAPITULASI BIAYA OPERASIONAL PT SO GOOD FOOD (FRESH) WH. CIKUPA"
    t.font = Font(name='Calibri', bold=True, size=11)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # Row 2: Periode
    ws.merge_cells('A2:T2')
    p = ws['A2']
    p.value = f"PERIODE : {m:02d} {BULAN[m]} {y}"
    p.font = Font(name='Calibri', bold=True, size=10)
    p.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # Row 3-5: Header
    HDR_BG = "BDD7EE"
    headers_r3 = [
        (1, "NO."), (2, "NAMA\nPENGGUNA"), (3, "HELPER"), (4, "NO. POLISI"),
        (5, "TANGGAL\nPENGGUNAAN"), (6, "JAM\nBRGKT"), (7, "JAM\nPLG"),
        (8, "KM\nAWAL"), (9, "KM\nAKHIR"),
    ]
    for col, txt in headers_r3:
        ws.merge_cells(start_row=3, start_column=col, end_row=5, end_column=col)
        cell(3, col, txt, bold=True, bg=HDR_BG)

    # RINCIAN BIAYA
    ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column=18)
    cell(3, 10, "RINCIAN BIAYA", bold=True, bg=HDR_BG)

    biaya_r4 = [
        "BBM (Rp.)", "TOL", "PARKIR", "PARKIR\nLIAR", "KULI /\nLAIN-LAIN",
        "NAMA HELPER\nHARIAN", "HELPER\nHARIAN", "TOTAL\nBIAYA", "RASIO\nBBM/LITER"
    ]
    for i, txt in enumerate(biaya_r4):
        col = 10 + i
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        cell(4, col, txt, bold=True, bg=HDR_BG)

    ws.row_dimensions[3].height = 32
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 12

    # Ambil harga BBM dari settings — tidak hardcode 12500
    _settings   = db.query(models.SystemSettings).first()
    _harga_bbm  = float(_settings.harga_bbm_per_liter or 12500.0) if _settings else 12500.0

    # Pre-load semua driver dan vehicle sekali — hindari N+1 query
    _all_drivers  = {d.driver_id: d for d in db.query(models.HRDriver).all()}
    _all_vehicles = {v.vehicle_id: v for v in db.query(models.FleetVehicle).all()}

    # Data rows
    DATA_START = 6
    for idx, e in enumerate(expenses, start=1):
        r = DATA_START + idx - 1
        helper_name = e.helper_name or ""

        drv          = _all_drivers.get(e.driver_id) if e.driver_id else None
        driver_name  = drv.name if drv else ""

        veh   = _all_vehicles.get(e.vehicle_id) if e.vehicle_id else None
        plate = veh.license_plate if veh else ""

        km_awal = e.km_awal or ""
        km_akhir = e.km_akhir or ""
        rasio = ""
        
        if e.km_awal and e.km_akhir and e.bbm and e.bbm > 0:
            jarak = e.km_akhir - e.km_awal
            liter = round(e.bbm / _harga_bbm, 1)
            if liter > 0:
                rasio = round(jarak / liter, 1)

        vals = [
            idx, driver_name, helper_name, plate,
            e.date.strftime("%d/%m/%Y") if e.date else "",
            e.jam_berangkat or "", e.jam_pulang or "",
            km_awal, km_akhir,
            e.bbm or 0, e.tol or 0, e.parkir or 0, e.parkir_liar or 0,
            (e.kuli_angkut or 0) + (e.lain_lain or 0),
            helper_name, e.lain_lain or 0,
            e.total or 0, rasio
        ]
        
        for ci, v in enumerate(vals, start=1):
            c2 = cell(r, ci, v, center=(ci not in [2, 3]))
            if 10 <= ci <= 18 and isinstance(v, (int, float)) and v:
                c2.number_format = '#,##0'

    # Grand Total
    total_row = DATA_START + len(expenses)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=9)
    cell(total_row, 1, "GRAND TOTAL", bold=True, bg="FFE699")
    
    total_cols = {
        10: 'bbm', 11: 'tol', 12: 'parkir', 13: 'parkir_liar',
        14: None, 15: None, 16: 'lain_lain', 17: 'total'
    }
    
    for col, field in total_cols.items():
        if field:
            total = sum(getattr(e, field) or 0 for e in expenses)
        elif col == 14:
            total = sum((e.kuli_angkut or 0) + (e.lain_lain or 0) for e in expenses)
        else:
            continue
        c2 = cell(total_row, col, total, bold=True, bg="FFE699")
        c2.number_format = '#,##0'

    # Tanda tangan
    sign_row = total_row + 3
    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=5)
    ws.merge_cells(start_row=sign_row, start_column=8, end_row=sign_row, end_column=13)
    ws.merge_cells(start_row=sign_row, start_column=15, end_row=sign_row, end_column=18)
    
    ws[f'A{sign_row}'] = "Dibuat oleh,"
    ws[f'H{sign_row}'] = "Mengetahui,"
    ws[f'O{sign_row}'] = "Disetujui"
    
    for r2 in [ws[f'A{sign_row}'], ws[f'H{sign_row}'], ws[f'O{sign_row}']]:
        r2.font = Font(name='Calibri', bold=True, size=9)

    # Column widths
    widths = [4, 22, 18, 13, 12, 7, 7, 8, 8, 12, 10, 10, 10, 12, 18, 12, 12, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"BOP_{y}_{m:02d}_{BULAN[m]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

@router.post("/bop-import-parse")
async def parse_bop_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    content = await file.read()
    buf = io.BytesIO(content)

    try:
        wb = load_workbook(buf, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File Excel tidak valid: {e}")

    # Header map yang sudah dibersihkan
    HEADER_MAP = {
        "NAMA PENGGUNA": "driver",
        "HELFER": "helper",
        "HELPER": "helper",
        "NO. POLISI": "plate",
        "NO.POLISI": "plate",
        "BBM": "bbm",
        "TOL": "tol",
        "PARKIR": "parkir",
        "PARKIR LIAR": "parkir_liar",
        "KULI / LAIN-LAIN": "kuli",
        "KULI ANGKUT": "kuli",
        "HELPER HARIAN": "helper_harian",
        "TOTAL BIAYA": "total",
        "JAM BRGKT": "jam_berangkat",
        "JAM PLG": "jam_pulang",
        "KM AWAL": "km_awal",
        "KM AKHIR": "km_akhir",
        "TANGGAL PENGGUNAAN": "tanggal",
        "NO.": "no",
    }

    col_map = {}
    data_start_row = None

    for row_idx in range(1, 9):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None: 
                continue
            
            # Bersihkan enter (\n) dan spasi berlebih dari cell Excel sebelum dicocokkan
            key = str(val).replace('\n', ' ').replace('\r', '').strip().upper()
            
            for hdr, field in HEADER_MAP.items():
                if key == hdr:
                    col_map[field] = col_idx
                    # Pakai max() agar header yang paling bawah yang menentukan baris data
                    # (BOP punya multi-row header: row 3 kolom utama, row 4 sub-header biaya)
                    data_start_row = max(data_start_row, row_idx + 1) if data_start_row else row_idx + 1

    if not col_map or not data_start_row:
        raise HTTPException(status_code=422, detail="Format Excel tidak dikenali. Pastikan header ada.")

    results = []
    warnings = []

    for row_idx in range(data_start_row, ws.max_row + 1):
        def g(field):
            if field not in col_map: return None
            return ws.cell(row=row_idx, column=col_map[field]).value

        no_val = g('no')
        if no_val is None: 
            continue
            
        try:
            no_int = int(str(no_val).strip())
        except ValueError:
            continue

        plate = str(g('plate') or "").strip().upper()
        driver = str(g('driver') or "").strip()
        helper = str(g('helper') or "").strip()

        def to_float(v):
            if v is None: return 0.0
            try: 
                return float(str(v).replace(',', '').replace('Rp', '').strip())
            except ValueError: 
                return 0.0

        bbm = to_float(g('bbm'))
        tol = to_float(g('tol'))
        parkir = to_float(g('parkir'))
        parkir_liar = to_float(g('parkir_liar'))
        kuli = to_float(g('kuli'))
        helper_h = to_float(g('helper_harian'))
        total = to_float(g('total')) or bbm + tol + parkir + parkir_liar + kuli + helper_h

        is_valid = True
        if plate:
            is_valid = db.query(models.FleetVehicle).filter(
                models.FleetVehicle.license_plate == plate
            ).first() is not None
            if not is_valid:
                warnings.append(f"Baris {no_int}: plat '{plate}' tidak ditemukan di master armada")

        results.append({
            "no": no_int,
            "plate": plate,
            "driver": driver,
            "helper": helper,
            "tanggal": str(g('tanggal') or date.today()),
            "jamBerangkat": str(g('jam_berangkat') or ""),
            "jamPulang": str(g('jam_pulang') or ""),
            "kmAwal": int(g('km_awal')) if g('km_awal') else None,
            "kmAkhir": int(g('km_akhir')) if g('km_akhir') else None,
            "bbm": bbm, "tol": tol, "parkir": parkir,
            "parkirLiar": parkir_liar, "kuliAngkut": kuli,
            "lainLain": helper_h, "helperName": helper,
            "total": total,
            "isValid": is_valid,
        })

    return {
        "status": "success",
        "rows": len(results),
        "warnings": warnings,
        "data": results,
    }

# ==========================================
# CRUD PENGELUARAN
# ==========================================
@router.post("/expenses", response_model=schemas.GenericResponse)
def create_expense(
    data: schemas.ExpenseCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin_distribusi", "manager_logistik", "kasir"))
):
    try:
        safe_vehicle_id = int(data.vehicle_id) if getattr(data, 'vehicle_id', None) else None
        safe_driver_id = int(data.driver_id) if getattr(data, 'driver_id', None) else None

        new_expense = models.OperationalExpense(
            id=data.id if getattr(data, 'id', None) else str(uuid.uuid4()),
            time=data.time,
            date=datetime.strptime(data.date, "%Y-%m-%d").date(),
            vehicle_id=safe_vehicle_id,
            driver_id=safe_driver_id,
            is_oncall=getattr(data, 'isOncall', False),
            bbm=getattr(data, 'bbm', 0.0) or 0.0,
            tol=getattr(data, 'tol', 0.0) or 0.0,
            parkir=getattr(data, 'parkir', 0.0) or 0.0,
            parkir_liar=getattr(data, 'parkirLiar', 0.0) or 0.0,
            kuli_angkut=getattr(data, 'kuliAngkut', 0.0) or 0.0,
            lain_lain=getattr(data, 'lainLain', 0.0) or 0.0,
            helper_name=getattr(data, 'helperName', ""),
            notes=getattr(data, 'notes', ""),
            total=getattr(data, 'total', 0.0) or 0.0,
            km_awal=getattr(data, 'kmAwal', None),
            km_akhir=getattr(data, 'kmAkhir', None),
            jam_berangkat=getattr(data, 'jamBerangkat', None),
            jam_pulang=getattr(data, 'jamPulang', None),
        )
        db.add(new_expense)
        db.commit()
        return {"status": "success", "message": "Biaya operasional berhasil dicatat."}
    except Exception as e:
        db.rollback() 
        logger.error(f"🚨 [CREATE EXPENSE] DB Error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Terjadi kesalahan internal server saat mencatat pengeluaran.")


@router.put("/expenses/{expense_id}", response_model=schemas.GenericResponse)
def update_expense(
    expense_id: str,
    data: schemas.ExpenseCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("admin_distribusi", "manager_logistik", "kasir"))
):
    expense = db.query(models.OperationalExpense).filter(models.OperationalExpense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="Data pengeluaran tidak ditemukan!")

    try:
        safe_vehicle_id = int(data.vehicle_id) if getattr(data, 'vehicle_id', None) else None
        safe_driver_id = int(data.driver_id) if getattr(data, 'driver_id', None) else None

        expense.time = data.time
        expense.date = datetime.strptime(data.date, "%Y-%m-%d").date()
        expense.vehicle_id = safe_vehicle_id
        expense.driver_id = safe_driver_id
        expense.is_oncall = getattr(data, 'isOncall', False)
        expense.bbm = getattr(data, 'bbm', 0.0) or 0.0
        expense.tol = getattr(data, 'tol', 0.0) or 0.0
        expense.parkir = getattr(data, 'parkir', 0.0) or 0.0
        expense.parkir_liar = getattr(data, 'parkirLiar', 0.0) or 0.0
        expense.kuli_angkut = getattr(data, 'kuliAngkut', 0.0) or 0.0
        expense.lain_lain = getattr(data, 'lainLain', 0.0) or 0.0
        expense.helper_name = getattr(data, 'helperName', "")
        expense.notes = getattr(data, 'notes', "")
        expense.total = getattr(data, 'total', 0.0) or 0.0
        expense.km_awal = getattr(data, 'kmAwal', None)
        expense.km_akhir = getattr(data, 'kmAkhir', None)
        expense.jam_berangkat = getattr(data, 'jamBerangkat', None)
        expense.jam_pulang = getattr(data, 'jamPulang', None)
        
        db.commit()
        return {"status": "success", "message": "Biaya operasional berhasil diupdate."}
    except Exception as e:
        db.rollback()
        logger.error(f"🚨 [UPDATE EXPENSE] DB Error untuk ID {expense_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Terjadi kesalahan internal server saat mengupdate pengeluaran.")


@router.get("/expenses", response_model=schemas.ExpenseListResponse)
def get_expenses(
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    query = db.query(models.OperationalExpense)
    
    if start_date:
        try:
            start_d = datetime.strptime(start_date, "%Y-%m-%d").date()
            query = query.filter(models.OperationalExpense.date >= start_d)
        except ValueError:
            raise HTTPException(status_code=400, detail="Format start_date harus YYYY-MM-DD")
            
    if end_date:
        try:
            end_d = datetime.strptime(end_date, "%Y-%m-%d").date()
            query = query.filter(models.OperationalExpense.date <= end_d)
        except ValueError:
            raise HTTPException(status_code=400, detail="Format end_date harus YYYY-MM-DD")

    expenses = query.order_by(models.OperationalExpense.created_at.desc()).all()
    results = [expense_to_dict(e) for e in expenses]
    return {"status": "success", "data": results}


@router.get("/expenses/today", response_model=schemas.ExpenseListResponse)
def get_today_expenses(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    today = date.today()
    expenses = db.query(models.OperationalExpense).filter(
        models.OperationalExpense.date == today
    ).order_by(models.OperationalExpense.created_at.desc()).all()
    
    results = [expense_to_dict(e) for e in expenses]
    return {"status": "success", "data": results}


@router.delete("/expenses/{expense_id}", response_model=schemas.GenericResponse)
def delete_expense(
    expense_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("manager_logistik", "admin_distribusi", "kasir"))
):
    expense = db.query(models.OperationalExpense).filter(models.OperationalExpense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="Data pengeluaran tidak ditemukan!")
        
    try:
        db.delete(expense)
        db.commit()
        return {"status": "success", "message": "Biaya operasional berhasil dihapus."}
    except Exception as e:
        db.rollback()
        logger.error(f"🚨 [DELETE EXPENSE] DB Error untuk ID {expense_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Terjadi kesalahan internal server saat menghapus pengeluaran.")